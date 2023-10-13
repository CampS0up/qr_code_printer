#!/usr/bin/env python3

import qrcode
from PIL import Image, ImageDraw, ImageFont
import time
import xlsxwriter
import os
import sys
import datetime
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def generate_qr_code(name_of_part, date, part_num, badge_number):
    # Create a QR code object
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=4,
    )

    # Add data to the QR code object
    qr.add_data(f"{name_of_part},{date},{part_num},{badge_number}")
    qr.make(fit=True)

    # Create an image object from the QR code object
    img = qr.make_image(fill_color="black", back_color="white")

    # Get the width and height of the image object
    width, height = img.size

    # Create a new image object with the size of the QR code and the name of the part
    new_img = Image.new("RGB", (width, height + 20), (255, 255, 255))

    # Paste the QR code image onto the new image object
    new_img.paste(img, (0, 20))

    # Create a text object with the name of the part.
    text = Image.new("RGB", (width, 35), (255, 255, 255))
    draw = ImageDraw.Draw(text)
    font = ImageFont.truetype("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", 24)
    draw.text(((width-len(name_of_part)*12)/2, 10), name_of_part, font=font, fill=(0, 0, 0))

    # Paste the text object onto the new image object
    new_img.paste(text, (0, 0))

    # Save the new image object as a PNG file
    new_img.save("qr_code.png")

def add_data_to_excel(name_of_part, date, part_number, badge_number):
    datasheetname = f"Data_for_{get_date('%m_%d_%Y')}.xlsx"
    datasheet_path = os.path.join("excel_data", datasheetname)

    date = get_date("%m-%d-%Y")

    # Create a new workbook if the file does not exist
    if not os.path.isfile(datasheet_path):
        workbook = Workbook()
        worksheet = workbook.active

        # Write the headers to the first row
        worksheet.cell(row=1, column=1, value='Name of Part')
        worksheet.cell(row=1, column=2, value='Date   ')
        worksheet.cell(row=1, column=3, value='Part Number')
        worksheet.cell(row=1, column=4, value='Badge Number')
        
        # Auto fit the columns
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Save the workbook
        workbook.save(datasheet_path)

    # Append the data to the existing workbook
    else:
        workbook = openpyxl.load_workbook(datasheet_path)
        worksheet = workbook.active

        # Get the next available row
        row = (worksheet.max_row + 1)

        # Write the data to the next available row
        worksheet.cell(row=row, column=1, value=name_of_part)
        worksheet.cell(row=row, column=2, value=date)
        worksheet.cell(row=row, column=3, value=part_number)
        worksheet.cell(row=row, column=4, value=badge_number)

        # Save the workbook
        workbook.save(datasheet_path)

def print_qr_code():
    os.system("lp -d 'Printer Name' qr_code.png")

    # Example usage
    #generate_qr_code("Part Name", time.strftime("%m.%d.%Y"), 1, 12345)
    #add_data_to_excel("Part Name", time.strftime("%m.%d.%Y"), 1, 12345)
    #print_qr_code()

def get_date(format="%m%d%Y"):
    # Get the current date and time
    current_date_time = datetime.now()
    # Format the date as 'mmddyyyy'
    formatted_date = current_date_time.strftime(format)
    return formatted_date

def generate_part_number():
    # Extract the sequence number from the last line
    try:
        # Open the log file in read mode
        with open('log.txt', 'r') as file:
            
            # Read the last line of the log file
            last_line = file.readlines()[-1]

        # Extract the sequence number from the last line
        if last_line.split()[-3] == get_date() or last_line.split()[-2] == get_shift():
            sequence_number = int(last_line.split()[-1]) + 1

        else:
            os.remove('log.txt')
            with open('log.txt', 'a') as file:
                file.write(f"This is the log file for {get_date('%m-%d-%Y')} \n")
                file.write("______________________________________________________________________\n")
            sequence_number = 1

    except IndexError:
        os.remove('log.txt')
        with open('log.txt', 'a') as file:
            file.write(f"This is the log file for {get_date('%m-%d-%Y')} \n")
            file.write("______________________________________________________________________\n")
        sequence_number=1

    # Generate the part number
    part_number = f"{get_date()} {get_shift()} {sequence_number}\n"
    
    # Open the log file in append mode
    with open('log.txt', 'a') as file:

        # Write the new part number to the log file
        file.write(f"{get_date()} {get_shift()} {sequence_number}\n")
    
    # Return the part number
    return (f"{get_date()}{get_shift()}{sequence_number}")

def get_shift():
  # Get the current time
  now = datetime.now()
  # Check if the current time is between 4:30 AM and 5 PM
  if now.hour >= 4 and now.hour < 17:
    return "D"
  # Otherwise, the current time must be between 5 PM and 4:30 AM
  else:
    return "N"

def main():
    if len(sys.argv) < 3:
        print("Usage: python qrcode.py generate|all")
        return

    command = sys.argv[1]
    name = sys.argv[2]
    date = get_date()
    part_num = generate_part_number()
    badge_number = 1
    shift = get_shift()
    if command == "generate":
        generate_qr_code(name, date, part_num, badge_number)
        print("QR code generated successfully.")
    elif command == "excel":
        add_data_to_excel(name, date, part_num, badge_number)
        print("Data added successfully.")
    elif command == "all":
        print("This command will do all the tasks.")
    else:
        print("Invalid command. Please use 'generate' or 'all'.")

if __name__ == "__main__":
    main()