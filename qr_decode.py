#!/usr/bin/env python3

import cv2 
import sys
from pyzbar.pyzbar import decode
import time
import xlsxwriter
import os
import sys
import datetime
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import re

def camera_decode(type):
    # Setup camera capture
    cap = cv2.VideoCapture()
    cap.open("http://10.183.44.176:8000/")
    cap.set(3, 640)  # set width
    cap.set(4, 480)  # set height
    done = 1
    # Check if argument is true
    if sys.argv[1] == "decode":
        while done == 1:
            # Read frame from camera
            success, img = cap.read()
            # Decode QR code or barcode
            for code in decode(img):
                if type == "part":
                    print(code.data.decode('utf-8'))
                    print("Code read input next code")
                    with open("part_num.txt", "a") as f:
                        f.write(f"{code.data.decode('utf-8')}\n")
                        done = 0
                if type == "badge":
                    print ("Employee ID: ")
                    employee_id = code.data.decode("utf-8")
                    print(f"{code.data.decode('utf-8')}\n")
                    print("Start scanning")
                    # Save the ID to a txt file
                    with open("employee_id.txt", "w") as f:
                        f.write(f"{employee_id}")
                        done = 0
            # Display image
            cv2.imshow('test', img)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
    else:
        # Read image file
        img = cv2.imread('qr_code.png')
        # Decode QR code or barcode
        try:
            codes = decode(img)
            for code in codes:
                type = code.type
                if type == "QRCODE":
                    print(code.data.decode('utf-8'))
                if type == "BARCODE":
                    print ("BAR")
        except:
            print("No QR code or barcode found")
    # Release camera
    cap.release()
    # Destroy all windows
    cv2.destroyAllWindows()

def get_date(format="%m%d%Y"):
    # Get the current date and time
    current_date_time = datetime.now()
    # Format the date as 'mmddyyyy'
    formatted_date = current_date_time.strftime(format)
    return formatted_date

def add_decoded_to_excel(name_of_assembly):
    datasheetname = f"Data_for_{get_date('%m_%d_%Y')}_{name_of_assembly}.xlsx"
    datasheet_path = os.path.join("excel_data", datasheetname)

    date = get_date("%m-%d-%Y")
    badge_number = get_badge_number()
    
    # Create a new workbook if the file does not exist
    if not os.path.isfile(datasheet_path):
        workbook = Workbook()
        worksheet = workbook.active

        # Write the headers to the first row
        worksheet.cell(row=1, column=1, value='Assembly Name')
        worksheet.cell(row=1, column=2, value='Date')
        worksheet.cell(row=1, column=3, value='Badge Number')
        worksheet.cell(row=1, column=4, value='Rear_profile')
        worksheet.cell(row=1, column=5, value='LH_Side_profile')
        worksheet.cell(row=1, column=6, value='RH_Side_profile')
        worksheet.cell(row=1, column=7, value='Canister_Assembly')
        worksheet.cell(row=1, column=8, value='Canister_and_Carpet_Assembly')
        worksheet.cell(row=1, column=9, value='Lid_Assembly')
        worksheet.cell(row=1, column=10, value='Part_7')
        worksheet.cell(row=1, column=11, value='Part_8')
        
        # Auto fit the columns
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        # Save the workbook
        workbook.save(datasheet_path)

    # Append the data to the existing workbook
    if os.path.isfile(datasheet_path):
        workbook = openpyxl.load_workbook(datasheet_path)
        worksheet = workbook.active

        # Get the next available row
        row = (worksheet.max_row + 1)
        worksheet.cell(row=row, column=1, value=name_of_assembly)
        worksheet.cell(row=row, column=2, value=date)
        worksheet.cell(row=row, column=3, value=badge_number)
        
        # Open and read the part_num.txt file
        with open("part_num.txt", "r") as f:
            lines = f.readlines()

        # Loop over each line in the file
        for line in lines:
            # Split the line into name, date, number, and badge
            name, date_unused, number, badge_unused = line.split(",")
            
            # Write the data to the next available row based on the name
            if name == "Rear_profile":
                worksheet.cell(row=row, column=4, value=number)
            elif name == "LH_Side_profile":
                worksheet.cell(row=row, column=5, value=number)
            elif name == "RH_Side_profile":
                worksheet.cell(row=row, column=6, value=number)
            elif name == "Canister_Assembly":
                worksheet.cell(row=row, column=7, value=number)
            elif name == "Canister_and_Carpet_Assembly":
                worksheet.cell(row=row, column=8, value=number)
            elif name == "Lid_Assembly":
                worksheet.cell(row=row, column=9, value=number)
            elif name == "Part_7":
                worksheet.cell(row=row, column=10, value=number)
            elif name == "Part_8":
                worksheet.cell(row=row, column=11, value=number)

        # Save the workbook
        workbook.save(datasheet_path)

def get_badge_number():
  with open("employee_id.txt", "r") as f:
    first_line = f.readline()
  return first_line.strip()

def clear_part_num_file():
    try:
        with open("part_num.txt", "w") as f:
            f.truncate(0)
        print("part_num.txt file has been cleared.")
    except FileNotFoundError:
        print("part_num.txt file not found.")

def main():
    if len(sys.argv) < 2:
        print("Usage: python qrcode.py generate|all")
        return

    command = sys.argv[1]
    function = sys.argv[2]
    name = sys.argv[2]
    if command == "decode":
        camera_decode(function)
        print("QR code generated successfully.")
    elif command == "excel":
        add_decoded_to_excel(name)
        print("Data added successfully.")
        clear_part_num_file()
    elif command == "all":
        print("This command will do all the tasks.")
    else:
        print("Invalid command. Please use 'generate' or 'all'.")

if __name__ == "__main__":
    main()
